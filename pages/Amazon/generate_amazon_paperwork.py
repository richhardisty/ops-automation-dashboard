"""
pages/Amazon/generate_amazon_paperwork.py

# ──────────────────────────────────────────────────────────────────────────────
# DEMO VERSION
# Sanitized for portfolio use. Credentials, internal file paths, company names,
# and email addresses have been replaced with placeholders. Requires a live
# NetSuite account, Amazon Vendor Central access, and a configured secrets.toml
# to run. See README.md for full context.
# ──────────────────────────────────────────────────────────────────────────────

Amazon VC Order Automation — plugs into the Ops Automation Streamlit dashboard.
Reads credentials from st.secrets (secrets.toml).
Imports login helpers from the Utilities/ folder.

Flow:
  1. Pre-flight checklist
  2. NetSuite: open pack worksheet URL → export combined CSV → wait for download
  3. Cross-reference check on the COMBINED file
     └─ If items missing → block, prompt for BPN + UPC, save, then continue
  4. Split combined CSV into per-PO files, delete combined
  5. For each PO (live progress cards):
       a. create_amazon_pack_sheet()  → XLSX + PDF
       b. amazon_shipment_builder()   → Ship Builder + labels
  6. Print pick tickets (NetSuite)
  7. Merge PDFs → zip → Outlook email draft
"""

import os, sys, time, glob, csv, shutil, threading, queue, math, gc, re, calendar, zipfile

import pandas as pd
import streamlit as st
from datetime import datetime, timedelta

# ── sys.path: Utilities (for amazon_login / netsuite_login) ─────────────────
UTILITIES_PATH = os.path.join(os.path.dirname(__file__), "..", "Utilities")
if UTILITIES_PATH not in sys.path:
    sys.path.insert(0, UTILITIES_PATH)

# ── Secrets ─────────────────────────────────────────────────────────────────
_s = st.secrets

DOWNLOAD_FOLDER                  = _s["DOWNLOAD_FOLDER"]
AMAZON_BASE_PROJECTS_FOLDER      = _s["AMAZON_BASE_PROJECTS_FOLDER"]
AMAZON_NETSUITE_PACK_WORKSHEET_US = _s["AMAZON_NETSUITE_PACK_WORKSHEET_US"]
AMAZON_SHIP_BUILDER              = _s["AMAZON_SHIP_BUILDER"]
NETSUITE_PICK_TICKETS            = _s["NETSUITE_PICK_TICKETS"]

# ── Derived paths ────────────────────────────────────────────────────────────
# Computed in a function to avoid triggering Windows shell/OneDrive integration
# on every Streamlit rerun (module-level os.path operations on OneDrive-synced
# paths can cause Explorer windows to open on RDP sessions).
AMAZON_AREA_FOLDER           = _s["AMAZON_AREA_FOLDER"]   # still used for Pack Worksheet / cross-ref
AMAZON_PACK_WORKSHEET_FOLDER = os.path.join(AMAZON_AREA_FOLDER, "Pack Worksheet")
AMAZON_PW_ITEM_CROSS_REF     = os.path.join(AMAZON_PACK_WORKSHEET_FOLDER,
                                             "Amazon Item Cross References.csv")
AMAZON_CARTON_WORKSHEET      = os.path.join(AMAZON_PACK_WORKSHEET_FOLDER,
                                             "Amazon Carton Worksheet.xlsx")

def _daily_paths():
    """Return today's output folder paths. Called only when actually needed."""
    now    = datetime.now()
    monday = now - timedelta(days=now.weekday())
    # Week folder:  Amazon Orders - Week of 2026-03 (Mar)-09
    week_folder = monday.strftime("Amazon Orders - Week of %Y-%m (%b)-%d")
    # Day subfolder: From 11-Mar
    day_folder  = now.strftime("From %d-%b")
    weekly = os.path.join(AMAZON_BASE_PROJECTS_FOLDER, week_folder)
    daily  = os.path.join(weekly, day_folder)
    return {
        "AMAZON_WEEKLY_FOLDER":       weekly,
        "AMAZON_DAILY_FOLDER":        daily,
        "AMAZON_DAILY_PACK_SHEETS":   os.path.join(daily, "Pack Sheets"),
        "AMAZON_DAILY_CARTON_LABELS": os.path.join(daily, "Carton Labels"),
        "AMAZON_DAILY_PALLET_LABELS": os.path.join(daily, "Pallet Labels"),
        "AMAZON_DAILY_LABEL_UPLOADS": os.path.join(daily, "Label Upload"),
    }

# Thin module-level references — pure string ops, no filesystem calls
_dp = _daily_paths()
AMAZON_WEEKLY_FOLDER       = _dp["AMAZON_WEEKLY_FOLDER"]
AMAZON_DAILY_FOLDER        = _dp["AMAZON_DAILY_FOLDER"]
AMAZON_DAILY_PACK_SHEETS   = _dp["AMAZON_DAILY_PACK_SHEETS"]
AMAZON_DAILY_CARTON_LABELS = _dp["AMAZON_DAILY_CARTON_LABELS"]
AMAZON_DAILY_PALLET_LABELS = _dp["AMAZON_DAILY_PALLET_LABELS"]
AMAZON_DAILY_LABEL_UPLOADS = _dp["AMAZON_DAILY_LABEL_UPLOADS"]


# ════════════════════════════════════════════════════════════════════════════
# SESSION-STATE INIT
# ════════════════════════════════════════════════════════════════════════════
def _init():
    defaults = {
        "ppr_phase":        "idle",
        # idle | preflight | netsuite_export | cross_ref_check | cross_ref_fix
        # | split_csv | processing | done
        "ppr_log":          [],
        "ppr_combined_path": None,
        "ppr_combined_df":  None,
        "ppr_cross_ref_df": None,
        "ppr_missing":      [],
        "ppr_missing_edits": {},
        "ppr_po_list":      [],
        "ppr_order_status": {},   # po -> {status, step, start, end}
        "ppr_running":      False,
        "ppr_status_q":     None,
        "ppr_run_thread":      None,
        "ppr_export_thread":   None,   # dedicated NetSuite export thread
        "ppr_export_q":        None,   # queue for export thread status
        "ppr_export_started":  False,  # guard so thread is only spawned once
        "ppr_run_started":      False,  # guard so _run_automation is only spawned once
        "ppr_onedrive_paused": False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init()


# ════════════════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════════════════
def _log(msg: str, level: str = "info"):
    ts = datetime.now().strftime("%H:%M:%S")
    st.session_state.ppr_log.append({"ts": ts, "msg": msg, "level": level})

def _render_log(container):
    level_colors = {"ok": "#3ee87a", "warn": "#f5c542", "error": "#f56060", "info": "#40aaff"}
    lines = []
    for e in reversed(st.session_state.ppr_log[-120:]):
        color = level_colors.get(e["level"], "#8b92a5")
        lines.append(
            f'<span style="color:#505568">[{e["ts"]}]</span> '
            f'<span style="color:{color}">{e["msg"]}</span>'
        )
    container.markdown(
        '<div style="background:#0a0c10;border:1px solid #1e2235;border-radius:8px;'
        'padding:12px 16px;font-family:monospace;font-size:0.76rem;line-height:1.8;'
        f'max-height:320px;overflow-y:auto">{"<br>".join(lines)}</div>',
        unsafe_allow_html=True,
    )

def _order_cards():
    status_icons = {"pending": "⏳", "running": "🔄", "complete": "✅", "failed": "❌"}
    colors = {
        "pending":  ("#1e2235", "#606880"),
        "running":  ("#0a2a4a", "#40aaff"),
        "complete": ("#0f3d26", "#3ee87a"),
        "failed":   ("#3d1010", "#f56060"),
    }
    html = ""
    for po in st.session_state.ppr_po_list:
        info   = st.session_state.ppr_order_status.get(po, {"status":"pending","step":"Queued"})
        st_key = info.get("status", "pending")
        bg, fg = colors.get(st_key, colors["pending"])
        icon   = status_icons.get(st_key, "⏳")
        step   = info.get("step", "")
        elapsed = ""
        if info.get("start"):
            end = info.get("end") or datetime.now()
            elapsed = f'{int((end - info["start"]).total_seconds())}s'
        html += (
            f'<div style="display:flex;align-items:center;gap:12px;'
            f'background:{bg};border:1px solid {fg}33;border-radius:8px;'
            f'padding:10px 16px;margin-bottom:8px;">'
            f'<div style="font-size:1.1rem">{icon}</div>'
            f'<div style="font-family:\'Courier New\',monospace;font-weight:700;'
            f'color:{fg};min-width:110px">PO {po}</div>'
            f'<div style="font-size:0.8rem;color:#8b92a5;flex:1">{step}</div>'
            f'<div style="font-size:0.75rem;color:#505568">{elapsed}</div>'
            f'</div>'
        )
    return html

def _find_combined_csv():
    matches = glob.glob(os.path.join(DOWNLOAD_FOLDER,
                                      "AmazonPackingWorksheetExportResults*.csv"))
    # Exclude already-split per-PO files (they contain " - ")
    combined = [p for p in matches if " - " not in os.path.basename(p)]
    return combined[0] if combined else None

def _do_cross_ref_check(combined_df):
    cross_df  = pd.read_csv(AMAZON_PW_ITEM_CROSS_REF, dtype={"Item": str, "UPC": str})
    ws_items  = set(combined_df["Item"].astype(str).str.strip().unique())
    ref_items = set(cross_df["Item"].astype(str).str.strip().unique())
    missing   = sorted(ws_items - ref_items, key=str)
    return cross_df, missing

def _save_cross_ref(base_df, new_rows):
    new_df = pd.DataFrame(new_rows)
    merged = pd.concat([base_df, new_df], ignore_index=True)
    merged = merged.drop_duplicates(subset=["Item"], keep="last")
    merged = merged.sort_values("Item", key=lambda x: x.astype(str))
    merged.to_csv(AMAZON_PW_ITEM_CROSS_REF, index=False, float_format="%g")
    return merged

def _split_combined(combined_df):
    po_list = []
    for po_raw in combined_df["PO"].unique():
        clean_po = str(po_raw).replace("PO:", "").strip()
        po_df    = combined_df[combined_df["PO"] == po_raw]
        out_path = os.path.join(
            DOWNLOAD_FOLDER,
            f"AmazonPackingWorksheetExportResults - {clean_po}.csv"
        )
        po_df.to_csv(out_path, index=False)
        po_list.append(clean_po)
        _log(f"Split → {os.path.basename(out_path)} ({len(po_df)} rows)", "ok")
    return sorted(po_list, key=str)

def _onedrive_exe_path():
    """
    Return OneDrive.exe path. Checks LOCALAPPDATA first (per diagnostic output),
    then scans for versioned subfolders, then falls back to Program Files.
    """
    localappdata = os.environ.get("LOCALAPPDATA", "")

    # 1. Direct path confirmed by registry on this machine
    direct = os.path.join(localappdata, "Microsoft", "OneDrive", "OneDrive.exe")
    if os.path.exists(direct):
        return direct

    # 2. Versioned subfolder (e.g. 26.032.0217.0003_1\OneDrive.exe)
    od_root = os.path.join(localappdata, "Microsoft", "OneDrive")
    if os.path.isdir(od_root):
        for entry in sorted(os.listdir(od_root), reverse=True):  # latest version first
            candidate = os.path.join(od_root, entry, "OneDrive.exe")
            if os.path.exists(candidate):
                return candidate

    # 3. System-wide installs
    for p in [
        r"C:\Program Files\Microsoft OneDrive\OneDrive.exe",
        r"C:\Program Files (x86)\Microsoft OneDrive\OneDrive.exe",
    ]:
        if os.path.exists(p):
            return p
    return None


def _is_onedrive_running():
    """Return True if any OneDrive-related sync process is running."""
    import subprocess
    r = subprocess.run(["tasklist", "/FO", "CSV", "/NH"],
                       capture_output=True, text=True)
    procs = r.stdout.lower()
    return any(p in procs for p in
               ("onedrive.exe", "onedrive.sync.service.exe", "filecoauth.exe"))


def _onedrive_pause():
    """
    Stops OneDrive sync completely by:
      1. Disabling the Startup scheduled task (stops it respawning on rerun)
      2. Killing all three sync processes: OneDrive.exe, OneDrive.Sync.Service.exe,
         FileCoAuth.exe  (diagnostic confirmed all three are running)
      3. Verifying they are gone
    Returns (success: bool, message: str).
    """
    import subprocess
    msgs = []

    # Step 1: Disable the OneDrive Startup Task so it cannot respawn during the run.
    # We re-enable it in _onedrive_resume(). Use a wildcard-free match on the
    # well-known task name prefix.
    try:
        r = subprocess.run(
            ["schtasks", "/query", "/FO", "CSV", "/NH"],
            capture_output=True, text=True
        )
        for line in r.stdout.splitlines():
            if "OneDrive Startup Task" in line:
                task_name = line.split(",")[0].strip().strip('"')
                subprocess.run(
                    ["schtasks", "/Change", "/TN", task_name, "/DISABLE"],
                    capture_output=True
                )
                msgs.append(f"Disabled scheduled task: {task_name}")
    except Exception as e:
        msgs.append(f"Could not disable scheduled task: {e}")

    # Step 2: Kill all three processes that keep OneDrive alive
    for proc_name in ("FileCoAuth.exe", "OneDrive.Sync.Service.exe", "OneDrive.exe"):
        try:
            subprocess.run(
                ["taskkill", "/F", "/IM", proc_name],
                capture_output=True
            )
            msgs.append(f"Killed {proc_name}")
        except Exception as e:
            msgs.append(f"Could not kill {proc_name}: {e}")

    # Step 3: Confirm they are gone (poll up to 8 s)
    for _ in range(8):
        time.sleep(1)
        if not _is_onedrive_running():
            msgs.append("All OneDrive processes stopped.")
            return True, " | ".join(msgs)

    # Still alive — something else is restarting it
    if _is_onedrive_running():
        return False, "OneDrive still running after kill attempts. " + " | ".join(msgs)
    return True, " | ".join(msgs)


def _onedrive_resume():
    """
    Re-enables the OneDrive Startup Task and restarts OneDrive.
    Returns (success: bool, message: str).
    """
    import subprocess
    msgs = []

    # Re-enable the scheduled task we disabled in _onedrive_pause()
    try:
        r = subprocess.run(
            ["schtasks", "/query", "/FO", "CSV", "/NH"],
            capture_output=True, text=True
        )
        for line in r.stdout.splitlines():
            if "OneDrive Startup Task" in line:
                task_name = line.split(",")[0].strip().strip('"')
                subprocess.run(
                    ["schtasks", "/Change", "/TN", task_name, "/ENABLE"],
                    capture_output=True
                )
                msgs.append(f"Re-enabled scheduled task: {task_name}")
    except Exception as e:
        msgs.append(f"Could not re-enable scheduled task: {e}")

    # Restart OneDrive with /background so it doesn't pop open a window
    exe = _onedrive_exe_path()
    if exe:
        try:
            subprocess.Popen([exe, "/background"])
            msgs.append("OneDrive restarted with /background.")
            return True, " | ".join(msgs)
        except Exception as e:
            msgs.append(f"Could not restart OneDrive.exe: {e}")
            return False, " | ".join(msgs)

    return False, "OneDrive.exe not found — please restart manually. " + " | ".join(msgs)


def _update_order(po, status, step):
    prev = st.session_state.ppr_order_status.get(po, {})
    if status == "running" and not prev.get("start"):
        prev["start"] = datetime.now()
    if status in ("complete", "failed"):
        prev["end"] = datetime.now()
    prev["status"] = status
    prev["step"]   = step
    st.session_state.ppr_order_status[po] = prev
    _log(f"PO {po}: {step}",
         "ok"    if status == "complete" else
         "error" if status == "failed"   else "info")


# ════════════════════════════════════════════════════════════════════════════
# NETSUITE EXPORT THREAD
# ════════════════════════════════════════════════════════════════════════════
def _run_netsuite_export(export_q: queue.Queue):
    """
    Opens Chrome, logs into NetSuite, clicks Export - CSV on the pack worksheet,
    waits for the file to land in Downloads, then signals completion.
    Sends ("status", level, message) tuples to export_q.
    Final signal is ("__DONE__", "ok"|"error", message).
    """
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service as ChromeService
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from webdriver_manager.chrome import ChromeDriverManager
        from netsuite_login import netsuite_login as _ns_login

        export_q.put(("status", "info", "Launching Chrome for NetSuite export…"))

        opts = webdriver.ChromeOptions()
        opts.add_experimental_option("prefs", {
            "download.default_directory": DOWNLOAD_FOLDER,
            "download.prompt_for_download": False,
            "profile.default_content_settings.popups": 0,
            "profile.content_settings.exceptions.automatic_downloads.*.setting": 1,
        })
        opts.add_argument("--disable-notifications")
        driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=opts,
        )

        def lm(msg):
            export_q.put(("status", "info", msg))

        export_q.put(("status", "info", "Navigating to NetSuite pack worksheet…"))
        driver.get(AMAZON_NETSUITE_PACK_WORKSHEET_US)
        driver.maximize_window()

        _ns_login(driver, lm)

        export_q.put(("status", "info", "Clicking Export - CSV…"))

        # Try clicking the Export button; handle the pop-up if it appears first
        def click_export():
            try:
                btn = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, '//div[@aria-label="Export - CSV"]'))
                )
                btn.click()
                return True
            except Exception:
                return False

        if not click_export():
            # Try closing a NetSuite pop-up that sometimes blocks it
            try:
                popup = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.ID, "uif10043"))
                )
                popup.click()
                export_q.put(("status", "info", "Closed NetSuite pop-up, retrying export…"))
                click_export()
            except Exception:
                pass

        export_q.put(("status", "info", "Waiting for CSV to download…"))

        # Poll Downloads folder until the file appears (up to 90 s)
        csv_pattern = os.path.join(DOWNLOAD_FOLDER, "AmazonPackingWorksheetExportResults*.csv")
        timeout, start = 90, time.time()
        found_path = None
        while time.time() - start < timeout:
            candidates = [
                p for p in glob.glob(csv_pattern)
                if " - " not in os.path.basename(p)       # not a split per-PO file
                and not p.endswith(".crdownload")          # not still downloading
            ]
            if candidates:
                found_path = candidates[0]
                break
            time.sleep(3)

        driver.quit()

        if found_path:
            export_q.put(("status", "ok",
                          f"CSV downloaded: {os.path.basename(found_path)}"))
            export_q.put(("__DONE__", "ok", found_path))
        else:
            export_q.put(("__DONE__", "error",
                          "Timed out waiting for CSV. Check NetSuite manually."))

    except Exception as e:
        try:
            driver.quit()
        except Exception:
            pass
        export_q.put(("__DONE__", "error", f"Export failed: {e}"))


# ════════════════════════════════════════════════════════════════════════════
# BACKGROUND WORKER
# ════════════════════════════════════════════════════════════════════════════
def _run_automation(po_list: list, status_q: queue.Queue):
    """
    Runs the full automation pipeline in a background thread.
    Sends (po_key, status, message) tuples to status_q.
    Special keys:  "__LOG__"  → log-only entry
                   "__ALL__"  → global status message
                   "__DONE__" → signals completion
    """
    # ── These run unconditionally before any imports that could fail ────────
    username = os.getlogin().lower()
    status_q.put(("__LOG__", "info", f"Running as: {username}"))

    # Create output folders before touching OneDrive paths
    for _d in [AMAZON_DAILY_FOLDER, AMAZON_DAILY_PACK_SHEETS,
               AMAZON_DAILY_CARTON_LABELS, AMAZON_DAILY_PALLET_LABELS,
               AMAZON_DAILY_LABEL_UPLOADS]:
        os.makedirs(_d, exist_ok=True)
    status_q.put(("__LOG__", "info", "Output folders ready."))

    # Pause OneDrive — must happen before any file writes into sync folders
    od_ok, od_msg = _onedrive_pause()
    status_q.put(("__LOG__", "ok" if od_ok else "warn", f"OneDrive: {od_msg}"))
    status_q.put(("__ONEDRIVE__", "paused", od_msg))

    try:
        # ── Windows-only imports ──────────────────────────────────────────
        import pythoncom
        pythoncom.CoInitialize()  # Required for COM (win32com) calls from a background thread
        import pyotp, openpyxl
        import win32com.client as win32
        import win32api, win32con
        from PyPDF2 import PdfMerger
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service as ChromeService
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from webdriver_manager.chrome import ChromeDriverManager

        # ── Login helpers from Utilities ──────────────────────────────
        from amazon_login   import amazon_login   as _amz_login
        from netsuite_login import netsuite_login as _ns_login

        # ── Automation log list ───────────────────────────────────────────
        _automation_log = []

        def log_message(msg):
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            _automation_log.append(f"{ts} - {msg}")
            status_q.put(("__LOG__", "info", msg))

        # ── Inline automation helpers (from amz_open_vc_automation.py) ────

        def wait_for_element(_drv, by, value, timeout=20):
            try:
                return WebDriverWait(_drv, timeout).until(EC.presence_of_element_located((by, value)))
            except Exception as e:
                log_message(f"Element not found: {e}"); return None

        def wait_for_element_to_be_clickable(_drv, by, value, timeout=30):
            try:
                return WebDriverWait(_drv, timeout).until(EC.element_to_be_clickable((by, value)))
            except Exception as e:
                log_message(f"Element not clickable: {e}"); return None

        def clear_input_field(_drv, element):
            _drv.execute_script("arguments[0].value = '';", element)

        def wait_for_file(file_prefix, directory, timeout=60):
            start_t = time.time()
            while time.time() - start_t < timeout:
                for f in os.listdir(directory):
                    if f.startswith(file_prefix):
                        return os.path.join(directory, f)
                time.sleep(2)
            raise TimeoutError(f"File '{file_prefix}' not found in '{directory}' within {timeout}s.")

        def label_find_latest_pdf(folder, prefix=""):
            for _r, _d, files in os.walk(folder):
                files = [f for f in files if f.endswith(".pdf") and f.startswith(prefix)]
                files.sort(key=lambda x: os.path.getctime(os.path.join(_r, x)), reverse=True)
                if files: return files[0], os.path.getctime(os.path.join(_r, files[0]))
            return None, None

        def process_labels(dl_folder, src_folder, tgt_folder, label_type, po, prefix=""):
            latest, _ = label_find_latest_pdf(dl_folder, prefix)
            if latest:
                new_name = f"{label_type} - {po}.pdf"
                os.rename(os.path.join(dl_folder, latest), os.path.join(dl_folder, new_name))
                shutil.move(os.path.join(dl_folder, new_name), os.path.join(tgt_folder, new_name))
            else:
                log_message(f"No PDF found for {label_type} (prefix={prefix!r})")

        def label_main(_usr, po):
            for folder in [AMAZON_DAILY_CARTON_LABELS, AMAZON_DAILY_PALLET_LABELS]:
                os.makedirs(folder, exist_ok=True)
            try:
                process_labels(DOWNLOAD_FOLDER, AMAZON_DAILY_LABEL_UPLOADS, AMAZON_DAILY_CARTON_LABELS, "Carton Labels", po, "cartonLabels")
                process_labels(DOWNLOAD_FOLDER, AMAZON_DAILY_LABEL_UPLOADS, AMAZON_DAILY_PALLET_LABELS, "Pallet Labels", po, "palletLabels")
            except Exception as e:
                log_message(f"label_main error: {e}")

        def merge_pdfs(src_folder, output_path):
            from PyPDF2 import PdfMerger as _PM
            merger = _PM()
            pdfs = sorted(glob.glob(os.path.join(src_folder, "*.pdf")))
            for p in pdfs: merger.append(p)
            if pdfs:
                merger.write(output_path); merger.close()
                log_message(f"Merged PDF: {os.path.basename(output_path)}")
            else:
                log_message(f"No PDFs to merge in {src_folder}")

        def zipdir(path, ziph, *ignore_folders):
            ignored = [os.path.join(path, f) for f in ignore_folders]
            for root, dirs, files in os.walk(path):
                if any(root.startswith(i) for i in ignored): continue
                for f in files:
                    if f.endswith(".zip") or f.endswith(".txt") or f.startswith("selected_dates"): continue
                    ziph.write(os.path.join(root, f), os.path.relpath(os.path.join(root, f), path))

        def send_email_with_outlook(subject, body, recipients, attachments=None):
            outlook = win32.Dispatch("outlook.application")
            mail = outlook.CreateItem(0)
            mail.Display()
            sig = mail.HTMLBody
            mail.To = ";".join(recipients)
            mail.Subject = subject
            mail.HTMLBody = body + sig
            if attachments:
                for att in attachments: mail.Attachments.Add(att)
            mail.Display(True)

        def select_date_for_po(_drv, po, excel_path):
            wb_sd = openpyxl.load_workbook(excel_path, data_only=True)
            sh_sd = wb_sd["Carton Breakdown"]
            so_sd   = sh_sd["A6"].value
            plts_sd = sh_sd["C3"].value or 0
            wb_sd.close()
            date_selected = None
            navigated_next = False

            def _next_month():
                btn = _drv.execute_script("""
                    let dp=document.querySelector("#builder-details-frd-datepicker").shadowRoot;
                    let cal=dp.querySelector("kat-calendar").shadowRoot;
                    return cal.querySelector("button.cal-rgt");
                """)
                if btn: _drv.execute_script("arguments[0].click();", btn); time.sleep(1.8); return True
                return False
            def _prev_month():
                btn = _drv.execute_script("""
                    let dp=document.querySelector("#builder-details-frd-datepicker").shadowRoot;
                    let cal=dp.querySelector("kat-calendar").shadowRoot;
                    return cal.querySelector("button.cal-lft");
                """)
                if btn: _drv.execute_script("arguments[0].click();", btn); time.sleep(1.8); return True
                return False
            def _day_cells():
                return _drv.execute_script("""
                    let dp=document.querySelector("#builder-details-frd-datepicker").shadowRoot;
                    let cal=dp.querySelector("kat-calendar").shadowRoot;
                    let t=cal.querySelector("table"); if(!t) return [];
                    return Array.from(t.querySelectorAll("td.day:not(.off):not(.disabled)"));
                """)
            def _day_info(td):
                return _drv.execute_script("""
                    let btn=arguments[0].querySelector("button.kat-no-style"); if(!btn) return null;
                    let day=parseInt(btn.getAttribute("data-day"),10);
                    let label=btn.getAttribute("aria-label")||"";
                    return {day, date_str:label.split(".")[0].trim(), is_green:btn.className.includes("color-code-07")};
                """, td)
            def _green_days(cells):
                g=[(c,_day_info(c)["day"],_day_info(c)["date_str"]) for c in cells if _day_info(c) and _day_info(c)["is_green"]]
                g.sort(key=lambda x:x[1], reverse=True); return g
            def _earliest(cells):
                cands=[(c,_day_info(c)["day"],_day_info(c)["date_str"]) for c in cells if _day_info(c)]
                if not cands: return None,None,None
                cands.sort(key=lambda x:x[1]); return cands[0]
            def _safe_click(el):
                try:
                    _drv.execute_script("arguments[0].scrollIntoView({block:'center'});",el); time.sleep(0.3)
                    _drv.execute_script("arguments[0].click();",el); time.sleep(0.8); return True
                except:
                    try: el.click(); time.sleep(0.8); return True
                    except: return False

            today_day = datetime.now().day
            if today_day >= 23:
                if _next_month():
                    navigated_next = True
                    greens = _green_days(_day_cells())
                    if greens:
                        td,_,ds = greens[0]
                        if _safe_click(td): date_selected = ds; log_message(f"Selected next-month green: {ds}")
            if date_selected is None:
                if navigated_next: _prev_month()
                cells = _day_cells(); greens = _green_days(cells)
                if greens:
                    td,_,ds = greens[0]
                    if _safe_click(td): date_selected = ds; log_message(f"Selected current-month green: {ds}")
                else:
                    td,_,ds = _earliest(cells)
                    if td and _safe_click(td): date_selected = ds; log_message(f"No greens → earliest: {ds}")
                    else: date_selected = "No selectable date found"
            if date_selected is None: date_selected = "Date not found"
            log_message(f"Date selected: {date_selected}")
            with open(os.path.join(AMAZON_DAILY_FOLDER,"selected_dates.csv"),"a",newline="") as f_sd:
                csv.writer(f_sd).writerow([so_sd, po, date_selected, "Parcel" if plts_sd==0 else "LTL"])
            return date_selected

        def create_amazon_pack_sheet(_usr, file_path):
            try:
                from openpyxl.utils.dataframe import dataframe_to_rows as _dtr
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter
                from openpyxl.worksheet.table import Table, TableStyleInfo
                destDir = AMAZON_PACK_WORKSHEET_FOLDER
                os.makedirs(destDir, exist_ok=True)
                df_ps = pd.read_csv(file_path).sort_values(by="Line #", ascending=False)
                po_ps = str(df_ps["PO"].iloc[0]).replace("PO:","").strip()
                log_message(f"Processing PO: {po_ps}")
                new_csv = os.path.join(destDir, f"Amazon Packing Worksheet - {po_ps}.csv")
                shutil.move(file_path, new_csv)
                t0 = time.time()
                while not os.path.exists(new_csv) and time.time()-t0 < 5: time.sleep(0.5)
                if not os.path.exists(new_csv): raise FileNotFoundError(f"CSV not found: {new_csv}")
                df_ps = pd.read_csv(new_csv).sort_values(by="Line #", ascending=False)
                out_xlsx = AMAZON_CARTON_WORKSHEET.replace(".xlsx", f" - {po_ps}.xlsx")

                WEIGHT_LIMIT, CUBE_LIMIT = 35, 5
                PALLET_WEIGHT_LIMIT, PALLET_CUBE_LIMIT = 1500, 70

                for idx_ps, row_ps in df_ps.iterrows():
                    sw = row_ps["Single Weight"] if row_ps["Single Weight"] != 0 else 0.01
                    sc = row_ps["Single Cube"]   if row_ps["Single Cube"]   != 0 else 0.01
                    if pd.isnull(row_ps["Master Cube"]):   df_ps.at[idx_ps,"Master Cube"]   = sc*row_ps["Master"]
                    if pd.isnull(row_ps["Master Weight"]): df_ps.at[idx_ps,"Master Weight"] = sw*row_ps["Master"]
                    if pd.isnull(row_ps["Inner Cube"]):    df_ps.at[idx_ps,"Inner Cube"]    = sc*row_ps["Inner"]
                    if pd.isnull(row_ps["Inner Weight"]):  df_ps.at[idx_ps,"Inner Weight"]  = sw*row_ps["Inner"]

                ctn_num = 1; cur_w = cur_c = 0; cur_items = []; mixed = []
                for _, row_ps in df_ps.iterrows():
                    for _ in range(int(row_ps["Master Packs"])):
                        mixed.append({"Ctn #":ctn_num,"Type":"Master","Item":row_ps["Item"],
                                      "Quantity":row_ps["Master"],"Line #":row_ps["Line #"],
                                      "Total Weight":row_ps["Master Weight"],"Total Cube":row_ps["Master Cube"]})
                        ctn_num += 1
                for _, row_ps in df_ps.iterrows():
                    for type_ in ["Inner","Single"]:
                        qty_ps = 1 if type_=="Single" else row_ps["Inner"]
                        for _ in range(int(row_ps[f"{type_} Packs"])):
                            if cur_w+row_ps[f"{type_} Weight"] > WEIGHT_LIMIT or cur_c+row_ps[f"{type_} Cube"] > CUBE_LIMIT:
                                if cur_items:
                                    mixed.append({"Ctn #":ctn_num,"Type":"Mixed","Items":cur_items.copy(),
                                                  "Line #":row_ps["Line #"],"Total Weight":cur_w,"Total Cube":cur_c})
                                    ctn_num += 1
                                cur_w = cur_c = 0; cur_items = []
                            cur_w += row_ps[f"{type_} Weight"]; cur_c += row_ps[f"{type_} Cube"]
                            cur_items.append({"Type":type_,"Item":row_ps["Item"],"Quantity":qty_ps,"Line #":row_ps["Line #"]})
                if cur_items:
                    mixed.append({"Ctn #":ctn_num,"Type":"Mixed","Items":cur_items.copy(),
                                  "Line #":cur_items[0]["Line #"],"Total Weight":cur_w,"Total Cube":cur_c})

                out_rows = []
                for item_ps in mixed:
                    if "Items" in item_ps:
                        for mi in sorted(item_ps["Items"], key=lambda x:(x["Type"],x["Item"])):
                            out_rows.append({"Ctn #":item_ps["Ctn #"],"Type":mi["Type"],"Item":mi["Item"],
                                             "Quantity":mi["Quantity"],"Line #":mi["Line #"],
                                             "Total Weight":item_ps["Total Weight"],"Total Cube":item_ps["Total Cube"]})
                    else: out_rows.append(item_ps)
                out_df = pd.DataFrame(out_rows)

                ctn_info = out_df.groupby("Ctn #").agg({"Total Weight":"first","Total Cube":"first"}).reset_index()
                total_ctns = len(ctn_info)
                if total_ctns <= 10:
                    out_df["Plt #"] = 0; plts_cnt = 0
                else:
                    total_cube_ps = df_ps["Total Cube"].sum()
                    plts_cnt = math.ceil(total_cube_ps / PALLET_CUBE_LIMIT)
                    cpc = math.ceil(total_ctns / plts_cnt)
                    plt_map = {ctn:(i//cpc)+1 for i,ctn in enumerate(ctn_info["Ctn #"])}
                    out_df["Plt #"] = out_df["Ctn #"].map(plt_map)

                cbd = out_df.groupby(["Plt #","Ctn #","Type","Item","Line #"]).agg(
                    Total_Quantity=("Quantity","sum"), Total_Weight=("Total Weight","first"),
                    Total_Cube=("Total Cube","first"), Count=("Type","size")).reset_index()
                cbd["Type x Count"] = cbd["Type"]+" x "+cbd["Count"].astype(str)
                cbd.columns = [c.replace("_"," ") for c in cbd.columns]
                cbd.rename(columns={"Total Quantity":"Quantity"},inplace=True)
                cbd["TypeOrder"] = cbd["Type"].map({"Master":0,"Inner":1,"Single":2})
                cbd = cbd.sort_values(["Plt #","Ctn #","TypeOrder","Line #"], ascending=[True,True,True,False])
                cbd = cbd[["Plt #","Ctn #","Line #","Item","Type x Count","Quantity","Total Weight","Total Cube"]]
                cbd.columns = cbd.columns.astype(str)
                cbd.rename(columns={"Quantity":"Qty","Total Weight":"Ctn Lbs","Total Cube":"Ctn Cube"},inplace=True)

                ctn_tots = cbd.groupby(["Plt #","Ctn #"]).agg({"Ctn Lbs":"first","Ctn Cube":"first"}).reset_index()
                plt_tots = ctn_tots.groupby("Plt #").agg({"Ctn Lbs":"sum","Ctn Cube":"sum"}).reset_index()
                cbd = cbd.sort_values(["Plt #","Ctn #"])
                cbd["IsFirstRowInCtn"] = ~cbd.duplicated(subset=["Plt #","Ctn #"])
                cbd = cbd.merge(plt_tots, on="Plt #", how="left", suffixes=("","_PltTotal"))
                cbd["IsFirstCtnInPlt"] = cbd["Ctn #"] == cbd.groupby("Plt #")["Ctn #"].transform("min")
                cbd["Plt Lbs"]  = cbd.apply(lambda r: r["Ctn Lbs_PltTotal"]  if r["IsFirstCtnInPlt"] and r["IsFirstRowInCtn"] else "",axis=1)
                cbd["Plt Cube"] = cbd.apply(lambda r: r["Ctn Cube_PltTotal"] if r["IsFirstCtnInPlt"] and r["IsFirstRowInCtn"] else "",axis=1)
                cbd["Ctn Lbs"]  = cbd.apply(lambda r: r["Ctn Lbs"]  if r["IsFirstRowInCtn"] else "",axis=1)
                cbd["Ctn Cube"] = cbd.apply(lambda r: r["Ctn Cube"] if r["IsFirstRowInCtn"] else "",axis=1)
                cbd.drop(columns=["Ctn Lbs_PltTotal","Ctn Cube_PltTotal","IsFirstRowInCtn","IsFirstCtnInPlt"],inplace=True)
                out_df.columns = out_df.columns.astype(str); cbd.columns = cbd.columns.astype(str)

                with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer_ps:
                    out_df.to_excel(writer_ps, sheet_name="Amazon Carton Worksheet", index=False)
                    wb_ps = writer_ps.book
                    ws_ps = wb_ps.create_sheet("Carton Breakdown")
                    ctn_count_ps = cbd["Ctn #"].nunique()
                    plts_count_ps = 0 if ctn_count_ps<=10 else math.ceil(df_ps["Total Cube"].sum()/PALLET_CUBE_LIMIT)
                    hcells = {
                        "A1":"Cartons","A3":ctn_count_ps,"C1":"Pallets","C3":plts_count_ps,
                        "E1":"Est. Total Weight","E3":math.ceil(df_ps["Total Weight"].sum()),
                        "H1":"Est. Total Cube","H3":math.ceil(df_ps["Total Cube"].sum()),
                        "A5":"Sales Order","A6":str(df_ps["SO"].iloc[0]).replace("SO:","").strip(),
                        "D5":"Purchase Order","D6":str(df_ps["PO"].iloc[0]).replace("PO:","").strip(),
                        "H5":"Ship Date","H6":str(df_ps["Ship Window Close"].max()) if "Ship Window Close" in df_ps.columns else "N/A",
                    }
                    hf_ps=Font(name="Helvetica",size=20,bold=True,underline="single")
                    shf_ps=Font(name="Helvetica",size=16,bold=True,underline="single")
                    shfs_ps=Font(name="Helvetica",size=14)
                    for cell_ps,val_ps in hcells.items():
                        ws_ps[cell_ps].value=val_ps
                        ws_ps[cell_ps].font=(shf_ps if cell_ps in("A5","D5","G5") else shfs_ps if cell_ps in("A6","D6","G6") else hf_ps)
                        ws_ps[cell_ps].alignment=Alignment(horizontal="center",vertical="center")
                    for cr in ["A1:B2","C1:D2","E1:G2","H1:J2","A3:B4","C3:D4","E3:G4","H3:J4","A5:C5","D5:G5","H5:J5","A6:C6","D6:G6","H6:J6"]:
                        ws_ps.merge_cells(cr)
                    fill_ps=PatternFill(start_color="D9D9D9",end_color="D9D9D9",fill_type="solid")
                    thick_ps=Border(left=Side(style="thick"),right=Side(style="thick"),top=Side(style="thick"),bottom=Side(style="thick"))
                    for row_ps2 in ws_ps.iter_rows(min_row=1,max_row=2,min_col=1,max_col=10): [setattr(c,"fill",fill_ps) for c in row_ps2]
                    for row_ps2 in ws_ps.iter_rows(min_row=5,max_row=5,min_col=1,max_col=10): [setattr(c,"fill",fill_ps) for c in row_ps2]
                    for row_ps2 in ws_ps.iter_rows(min_row=1,max_row=6,min_col=1,max_col=10): [setattr(c,"border",thick_ps) for c in row_ps2]
                    end_cl=get_column_letter(10); end_rw=8+len(cbd)
                    for _rn2, row_ps2 in enumerate(ws_ps.iter_rows(min_row=8,max_row=end_rw,min_col=1,max_col=10),8):
                        for cell_ps2 in row_ps2:
                            cl_ps=cell_ps2.coordinate
                            if cl_ps=="A8": cell_ps2.border=Border(left=Side(style="thick"),top=Side(style="thick"))
                            elif cl_ps in [f"{l}8" for l in "BCDEFGHI"]: cell_ps2.border=Border(top=Side(style="thick"))
                            elif cl_ps=="J8": cell_ps2.border=Border(top=Side(style="thick"),right=Side(style="thick"))
                            elif cl_ps==f"A{end_rw}": cell_ps2.border=Border(left=Side(style="thick"),bottom=Side(style="thick"))
                            elif cl_ps==f"{end_cl}{end_rw}": cell_ps2.border=Border(right=Side(style="thick"),bottom=Side(style="thick"))
                            elif cell_ps2.column==1: cell_ps2.border=Border(left=Side(style="thick"))
                            elif cell_ps2.column==10: cell_ps2.border=Border(right=Side(style="thick"))
                            elif 1<cell_ps2.column<10 and cl_ps.endswith(str(end_rw)): cell_ps2.border=Border(bottom=Side(style="thick"))
                    for ri_ps,row_ps2 in enumerate(_dtr(cbd,index=False,header=True),8):
                        for ci_ps,val_ps in enumerate(row_ps2,1): ws_ps.cell(row=ri_ps,column=ci_ps,value=val_ps)
                    for col_l_ps,w_ps in zip("ABCDEFGHIJ",[8,8,8,12,12,8,10,10,12,12]):
                        ws_ps.column_dimensions[col_l_ps].width=w_ps
                    for row_ps2 in ws_ps.iter_rows():
                        for c in row_ps2:
                            cl2_ps=getattr(c,"column_letter",None)
                            c.font=Font(name="Helvetica",size=11,bold=(cl2_ps=="A" or c.row==8))
                            c.alignment=openpyxl.styles.Alignment(horizontal="center",vertical="center")
                    for c in ws_ps["A"]: c.font=Font(name="Helvetica",size=14,bold=True)
                    for c in ws_ps["B"]: c.font=Font(name="Helvetica",size=14,bold=True)
                    for c in ws_ps["1"]: c.font=Font(name="Helvetica",size=16,bold=True,underline="single"); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
                    for c in ws_ps["3"]: c.font=Font(name="Helvetica",size=16)
                    for c in ws_ps["5"]: c.font=Font(name="Helvetica",size=16,bold=True,underline="single")
                    for c in ws_ps["6"]: c.font=Font(name="Helvetica",size=14)
                    for c in ws_ps["8"]: c.font=Font(name="Helvetica",size=10,bold=True)
                    ws_ps.row_dimensions[1].height=20; ws_ps.row_dimensions[2].height=20; ws_ps.row_dimensions[5].height=30
                    ec_ps=get_column_letter(ws_ps.max_column); er_ps=ws_ps.max_row

                    # Table with color stripes (TableStyleMedium21 = blue/white alternating)
                    tbl=Table(displayName="CartonBreakdown", ref=f"A8:{ec_ps}{er_ps}")
                    tbl.tableStyleInfo=TableStyleInfo(
                        name="TableStyleMedium21", showFirstColumn=False, showLastColumn=False,
                        showRowStripes=True, showColumnStripes=False)
                    ws_ps.add_table(tbl)

                    ws_ps.print_area=f"A1:{ec_ps}{er_ps}"
                    # Narrow margins so all 10 columns fit on one page width
                    ws_ps.page_margins.left=0.25
                    ws_ps.page_margins.right=0.25
                    ws_ps.page_margins.top=0.5
                    ws_ps.page_margins.bottom=0.5
                    ws_ps.page_margins.header=0.3
                    ws_ps.page_margins.footer=0.3
                    ws_ps.page_setup.orientation=ws_ps.ORIENTATION_PORTRAIT
                    ws_ps.page_setup.fitToWidth=1
                    ws_ps.page_setup.fitToHeight=0
                    ws_ps.sheet_properties.pageSetUpPr.fitToPage=True

                    # Print Carton Label Worksheet tab
                    if "Print Carton Label Worksheet" in wb_ps.sheetnames:
                        del wb_ps["Print Carton Label Worksheet"]
                    pivot_ws=wb_ps.create_sheet("Print Carton Label Worksheet")
                    max_plt_ps=out_df["Plt #"].max()
                    if max_plt_ps==0:
                        hdrs=["Carton label barcode","PO","SKU - ASIN","Units per carton","Expiration date","Manufacture date","Lot number"]
                    else:
                        hdrs=["Pallet label barcode","Carton label barcode","PO","SKU - ASIN","Units per carton","Expiration date","Manufacture date","Lot number"]
                    for ci_ps,h in enumerate(hdrs,1): pivot_ws.cell(row=1,column=ci_ps,value=h)
                    special_date_ps=datetime(2026,12,31)
                    special_items_ps=[str(i) for i in [
                        "30407","30583","30588","30415","30832","20818","30376","30321","30803","21498",
                        "30399","30381","30325","20865","19279","30323","30324","30899","20755","20617",
                        "21532","19266","30685","30896","20766","20765",
                        30407,30583,30588,30415,30832,20818,30376,30321,30803,21498,30399,30381,30325,
                        20865,19279,30323,30324,30899,20755,20617,21532,19266,30685,30896,20766,20765]]
                    grp=out_df.groupby(["Plt #","Ctn #","Item","Line #"]).agg(
                        Total_Quantity=("Quantity","sum")).reset_index().sort_values(["Plt #","Ctn #","Line #"], ascending=[True,True,False])
                    r_ps=2
                    for _,row_ps3 in grp.iterrows():
                        exp_ps=special_date_ps.strftime("%m/%d/%Y") if str(row_ps3["Item"]) in special_items_ps else ""
                        if max_plt_ps>0:
                            pivot_ws.cell(r_ps,1,row_ps3["Plt #"]); pivot_ws.cell(r_ps,2,row_ps3["Ctn #"])
                            pivot_ws.cell(r_ps,3,df_ps["PO"].iloc[0]); pivot_ws.cell(r_ps,4,row_ps3["Item"])
                            pivot_ws.cell(r_ps,5,row_ps3["Total_Quantity"]); pivot_ws.cell(r_ps,6,exp_ps)
                            pivot_ws.cell(r_ps,7,""); pivot_ws.cell(r_ps,8,"")
                        else:
                            pivot_ws.cell(r_ps,1,row_ps3["Ctn #"]); pivot_ws.cell(r_ps,2,df_ps["PO"].iloc[0])
                            pivot_ws.cell(r_ps,3,row_ps3["Item"]); pivot_ws.cell(r_ps,4,row_ps3["Total_Quantity"])
                            pivot_ws.cell(r_ps,5,exp_ps); pivot_ws.cell(r_ps,6,""); pivot_ws.cell(r_ps,7,"")
                        r_ps+=1
                    for col_ps in pivot_ws.columns:
                        mx_ps=max((len(str(c.value)) for c in col_ps if c.value),default=0)
                        pivot_ws.column_dimensions[col_ps[0].column_letter].width=mx_ps+10
                    cref_df_ps=pd.read_csv(AMAZON_PW_ITEM_CROSS_REF)
                    cref_df_ps["Item"]=cref_df_ps["Item"].astype(str).str.strip().str.upper()
                    cref_dict_ps=cref_df_ps.set_index("Item")["BPN"].to_dict()
                    for row_ps4 in pivot_ws.iter_rows(min_row=2,max_row=pivot_ws.max_row):
                        ac=row_ps4[hdrs.index("SKU - ASIN")]
                        av=str(ac.value).strip().upper()
                        if av in cref_dict_ps: ac.value=cref_dict_ps[av]

                t0=time.time()
                while not os.path.exists(out_xlsx) and time.time()-t0<5: time.sleep(0.5)
                if not os.path.exists(out_xlsx): raise FileNotFoundError(f"XLSX not found: {out_xlsx}")
                log_message(f"XLSX saved: {os.path.basename(out_xlsx)}")

                excel_app_ps=win32.Dispatch("Excel.Application")
                wb_win=excel_app_ps.Workbooks.Open(out_xlsx)
                psn=wb_win.Sheets("Carton Breakdown").Range("D6").Value
                os.makedirs(AMAZON_DAILY_PACK_SHEETS,exist_ok=True)
                pdf_ps=os.path.join(AMAZON_DAILY_PACK_SHEETS,f"Pack Sheet - {psn}.pdf")
                try:
                    excel_app_ps.DisplayAlerts=False
                    wb_win.Sheets("Carton Breakdown").ExportAsFixedFormat(0,pdf_ps)
                    log_message(f"Pack Sheet - {psn} generated")
                    wb_win.Close(SaveChanges=False); excel_app_ps.Quit()
                except Exception as e:
                    log_message(f"PDF export error: {e}")
                    try: wb_win.Close(SaveChanges=False); excel_app_ps.Quit()
                    except Exception: pass
                finally:
                    gc.collect()
                return out_xlsx
            except Exception as e:
                log_message(f"create_amazon_pack_sheet error: {e}")
                return None

        def open_amazon_ship_builder(_drv):
            _drv.get(AMAZON_SHIP_BUILDER); _drv.maximize_window()
            log_message("Opened Amazon Shipment Builder"); time.sleep(2)

        def amazon_shipment_builder(_drv, _usr, excel_path):
            wb_sb=openpyxl.load_workbook(excel_path,data_only=True); sh_sb=wb_sb["Carton Breakdown"]
            po_sb=sh_sb["D6"].value; so_sb=sh_sb["A6"].value
            ctns_sb=sh_sb["A3"].value or 0; plts_sb=sh_sb["C3"].value or 0
            wt_sb=sh_sb["E3"].value or 0; cb_sb=sh_sb["H3"].value or 0
            wb_sb.close()
            log_message(f"Shipment Builder: PO={po_sb} Cartons={ctns_sb} Pallets={plts_sb}")
            ea2=win32.Dispatch("Excel.Application"); ea2.Workbooks.Open(excel_path)
            ea2.Workbooks(1).Close(SaveChanges=0); ea2.Application.Quit(); time.sleep(2)

            sf_sb=wait_for_element_to_be_clickable(_drv,By.ID,"sb-table-search",120)
            if sf_sb: sf_sb.click(); sf_sb.send_keys(po_sb); sf_sb.send_keys(Keys.RETURN); log_message(f"Searched PO: {po_sb}")
            time.sleep(2)
            ab=wait_for_element_to_be_clickable(_drv,By.ID,f"addButton-{po_sb}",120)
            if ab: ab.click(); log_message("Added PO to shipment")
            time.sleep(2)
            c2=wait_for_element_to_be_clickable(_drv,By.XPATH,'//kat-button[@class="sb-footer-button" and @label="Continue to step 2"]',120)
            if c2: c2.click(); log_message("Continue to step 2")
            time.sleep(10)
            for el_id_sb,val_sb,lbl_sb in [
                ("packing-detail-carton-count",ctns_sb,"Cartons"),
                ("packing-detail-pallet-count",plts_sb,"Pallets"),
                ("packing-detail-weight",wt_sb,"Weight"),
                ("packing-detail-volume",cb_sb,"Cube"),
            ]:
                el_sb=wait_for_element_to_be_clickable(_drv,By.ID,el_id_sb)
                if el_sb: el_sb.click(); clear_input_field(_drv,el_sb); el_sb.send_keys(val_sb); log_message(f"{lbl_sb}: {val_sb}")
            sd_sb=wait_for_element_to_be_clickable(_drv,By.ID,"packing-detail-pallet-type")
            if sd_sb:
                sd_sb.click(); time.sleep(0.5)
                res_sb=_drv.execute_script("""
                    let dd=document.querySelector("#packing-detail-pallet-type").shadowRoot;
                    let opt=dd.querySelector("div>div:nth-child(3)>div>div>div>slot:nth-child(2)>kat-option:nth-child(2)");
                    if(opt){opt.click();return true;}return false;
                """)
                log_message("Set Unstackable" if res_sb else "Unstackable not found")
            so_inp_sb=wait_for_element_to_be_clickable(_drv,By.ID,"shipment-reference-input")
            if so_inp_sb: so_inp_sb.click(); clear_input_field(_drv,so_inp_sb); so_inp_sb.send_keys(so_sb); log_message(f"SO: {so_sb}")
            c3=wait_for_element_to_be_clickable(_drv,By.XPATH,'//kat-button[@class="sb-footer-button" and @label="Continue to step 3"]')
            if c3:
                for _att in range(2):
                    try: c3.click(); log_message("Continue to step 3"); break
                    except Exception as e_sb:
                        if "element click intercepted" in str(e_sb).lower():
                            _drv.execute_script("window.scrollTo(0,document.body.scrollHeight);"); time.sleep(2)
                        else: raise
            time.sleep(2)
            dp_sb=wait_for_element_to_be_clickable(_drv,By.XPATH,'//kat-date-picker[@id="builder-details-frd-datepicker"]',120)
            if dp_sb: dp_sb.click(); log_message("Clicked date picker")
            else: log_message("Date picker not found"); _drv.quit(); return
            time.sleep(2)
            select_date_for_po(_drv,po_sb,excel_path)
            time.sleep(1)
            cfm=wait_for_element_to_be_clickable(_drv,By.XPATH,'//kat-button[@label="Confirm and submit shipment"]')
            if cfm:
                for _att in range(2):
                    try: cfm.click(); log_message("Confirmed and submitted shipment"); break
                    except Exception as e_sb:
                        if "element click intercepted" in str(e_sb).lower():
                            _drv.execute_script("window.scrollTo(0,document.body.scrollHeight);"); time.sleep(2)
                        else: raise
            time.sleep(5)
            log_message("Extracting Shipment ID...")
            se=WebDriverWait(_drv,120).until(EC.presence_of_element_located((By.CSS_SELECTOR,
                "#pickup-details-div > div.kat-row.sb-details-row > div.kat-col-md-8 > h4")))
            arn_sb=se.text.replace("Shipment ID (ARN): ",""); log_message(f"ARN: {arn_sb}"); time.sleep(5)
            _drv.get(f"https://vendorcentral.amazon.com/kt/vendor/members/afi-shipment-mgr/asnsubmission?arn={arn_sb}&asnId=&isLabelMapping=true")
            time.sleep(5)
            sbjs="""
                let sb2=document.querySelector('kat-button.label-footer-button[label="Continue to step 2"]');
                if(sb2){let ib=sb2.shadowRoot.querySelector("button");if(ib){ib.click();return true;}}return false;
            """
            try: WebDriverWait(_drv,20).until(lambda d:d.execute_script(sbjs)); log_message("Continue to step 2 (shadow DOM)")
            except Exception:
                try:
                    fb2=WebDriverWait(_drv,10).until(EC.element_to_be_clickable((By.XPATH,'//kat-button[@class="label-footer-button" and @label="Continue to step 2"]')))
                    _drv.execute_script("arguments[0].shadowRoot.querySelector('button').click();",fb2)
                    log_message("Continue to step 2 (XPath fallback)")
                except Exception as e_sb: log_message(f"Continue step 2 failed: {e_sb}"); raise
            time.sleep(3)
            if plts_sb>0:
                try:
                    prg2=WebDriverWait(_drv,10).until(EC.presence_of_element_located((By.CSS_SELECTOR,"div.label-mapped-items-labeltype kat-radiobutton-group")))
                    prg2.find_element(By.CSS_SELECTOR,"kat-radiobutton:first-child span").click(); log_message("AMZNCC pallet labels")
                except Exception as e_sb: log_message(f"Pallet label type error: {e_sb}")
                time.sleep(2)
                try:
                    dlb=WebDriverWait(_drv,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"div.label-mapped-items-labeltype kat-button")))
                    _drv.execute_script("arguments[0].shadowRoot.querySelector('button').click();",dlb); log_message("Downloading AMZNCC labels")
                except Exception as e_sb: log_message(f"AMZNCC download error: {e_sb}")
                time.sleep(2)
                try:
                    cs3=WebDriverWait(_drv,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#label-items-footer kat-popover kat-button")))
                    _drv.execute_script("arguments[0].shadowRoot.querySelector('button').click();",cs3); log_message("Continue to step 3")
                except Exception as e_sb: log_message(f"Step 3 error: {e_sb}")
            else:
                try:
                    dlb2=WebDriverWait(_drv,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"div.label-mapped-items-labeltype kat-button")))
                    _drv.execute_script("arguments[0].shadowRoot.querySelector('button').click();",dlb2); log_message("Downloading carton labels")
                except Exception as e_sb: log_message(f"Download error: {e_sb}")
                time.sleep(2)
                try:
                    cs3b=WebDriverWait(_drv,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#label-items-footer kat-popover kat-button")))
                    _drv.execute_script("arguments[0].shadowRoot.querySelector('button').click();",cs3b); log_message("Continue to step 3")
                except Exception as e_sb: log_message(f"Step 3 error: {e_sb}")
            time.sleep(2)
            ldf=pd.read_excel(excel_path, sheet_name="Print Carton Label Worksheet")
            amzncc_p=os.path.join(DOWNLOAD_FOLDER,"AMZNCCLabels.csv")
            os.makedirs(AMAZON_DAILY_LABEL_UPLOADS,exist_ok=True)
            ul_file=os.path.join(AMAZON_DAILY_LABEL_UPLOADS,f"Label Upload - {po_sb}.csv")
            if os.path.exists(amzncc_p): adf=pd.read_csv(amzncc_p,header=None)
            else: raise FileNotFoundError(f"{amzncc_p} does not exist")
            def _rep_bc(df_rb,col_rb,adf_rb):
                rows_del=set()
                if col_rb not in df_rb.columns: return rows_del
                def _rep(bc):
                    try:
                        rn=int(bc)-1
                        if 0<=rn<len(adf_rb): rows_del.add(rn); return adf_rb.iloc[rn,0]
                    except ValueError: return bc
                df_rb[col_rb]=df_rb[col_rb].apply(_rep); return rows_del
            crd=_rep_bc(ldf,"Carton label barcode",adf); adf=adf.drop(list(crd),errors="ignore")
            if adf.empty: os.remove(amzncc_p)
            else:
                adf.to_csv(amzncc_p,header=False,index=False); adf=pd.read_csv(amzncc_p,header=None)
                if not adf.empty and "Pallet label barcode" in ldf.columns:
                    prd=_rep_bc(ldf,"Pallet label barcode",adf); adf=adf.drop(list(prd),errors="ignore")
                    if adf.empty: os.remove(amzncc_p)
                    else: adf.to_csv(amzncc_p,header=False,index=False)
            if "SKU - ASIN" in ldf.columns:
                ldf["SKU - ASIN"]=ldf["SKU - ASIN"].apply(lambda x:re.sub(r"[^A-Za-z0-9]","",str(x)))
            ldf.to_csv(ul_file,index=False); log_message(f"Label upload CSV: {ul_file}")
            _drv.execute_script("""
                const fi=document.querySelector(
                  "#asn-websheet-import-export-buttons>div>div>div>div.kat-col-lg-6>kat-file-upload"
                ).shadowRoot.querySelector("input[type='file']");
                fi.style.display="block"; return fi;
            """).send_keys(ul_file)
            log_message("File selected for upload"); time.sleep(2)
            _drv.execute_script("window.scrollTo(0,document.body.scrollHeight);"); time.sleep(2)
            _drv.execute_script("""
                document.querySelector("#label-items-footer>div>kat-button").shadowRoot.querySelector("button").click();
            """); log_message("Clicked Save as draft"); time.sleep(5)
            _drv.execute_script("""
                const pb=document.querySelector("#label-items-footer>div>kat-button:nth-child(1)").shadowRoot.querySelector("button");
                pb.scrollIntoView(true); pb.click();
            """); log_message("Clicked Print Labels"); time.sleep(2)
            if plts_sb>0:
                plb=_drv.execute_script("""
                    return document.querySelector(
                        "#center-app-div>div:nth-child(5)>div:nth-child(6)>kat-modal>div:nth-child(6)>kat-button:nth-child(1)"
                    ).shadowRoot.querySelector("button");
                """); _drv.execute_script("arguments[0].click();",plb); log_message("Printing Pallet Labels")
                try: wait_for_file("palletLabels",DOWNLOAD_FOLDER)
                except TimeoutError as e_sb: log_message(f"Pallet labels timeout: {e_sb}"); raise
                time.sleep(5)
                clb=_drv.execute_script("""
                    return document.querySelector(
                        "#center-app-div>div:nth-child(5)>div:nth-child(6)>kat-modal>div:nth-child(6)>kat-button:nth-child(2)"
                    ).shadowRoot.querySelector("button");
                """); _drv.execute_script("arguments[0].click();",clb); log_message("Printing Carton Labels")
                try: wait_for_file("cartonLabels",DOWNLOAD_FOLDER)
                except TimeoutError as e_sb: log_message(f"Carton labels timeout: {e_sb}"); raise
            else:
                _drv.execute_script("""
                    document.querySelector(
                        "#center-app-div>div:nth-child(5)>div:nth-child(6)>kat-modal>div:nth-child(6)>kat-button"
                    ).shadowRoot.querySelector("button").click();
                """); log_message("Printing Carton Labels only")
                try: wait_for_file("cartonLabels",DOWNLOAD_FOLDER)
                except TimeoutError as e_sb: log_message(f"Carton labels timeout: {e_sb}"); raise
            time.sleep(5); label_main(_usr,po_sb); time.sleep(1)

        # ── Close stray Excel instances ───────────────────────────────────
        try:
            for proc in win32.GetObject("winmgmts:").InstancesOf("Win32_Process"):
                if proc.Name == "EXCEL.EXE":
                    handle = int(proc.Handle)
                    win32api.TerminateProcess(
                        win32api.OpenProcess(win32con.PROCESS_TERMINATE, 0, handle), 0
                    )
        except Exception:
            pass

        # ── Chrome (kiosk-print for label PDFs) ───────────────────────────
        opts = webdriver.ChromeOptions()
        opts.add_experimental_option("prefs", {
            "download.default_directory": DOWNLOAD_FOLDER,
            "download.prompt_for_download": False,
            "plugins.always_open_pdf_externally": True,
            "profile.default_content_settings.popups": 0,
            "printing.print_preview_sticky_settings.appState": (
                '{"recentDestinations":[{"id":"Save as PDF","origin":"local"}],'
                '"selectedDestinationId":"Save as PDF","version":2}'
            ),
        })
        opts.add_argument("--kiosk-printing")
        opts.add_argument("--disable-notifications")
        driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=opts,
        )

        # Wrap log_message so automation helpers write back to status_q
        def lm(msg):
            status_q.put(("__LOG__", "info", msg))

        # ── Per-PO loop ────────────────────────────────────────────────────
        login_done = False
        csv_files  = sorted(glob.glob(
            os.path.join(DOWNLOAD_FOLDER,
                         "AmazonPackingWorksheetExportResults - *.csv")
        ))

        for file_path in csv_files:
            po = (os.path.basename(file_path)
                  .replace("AmazonPackingWorksheetExportResults - ", "")
                  .replace(".csv", ""))
            try:
                status_q.put((po, "running", "Creating pack sheet…"))
                excel_path = create_amazon_pack_sheet(username, file_path)
                if not excel_path:
                    status_q.put((po, "failed", "Pack sheet creation failed"))
                    continue

                if not login_done:
                    status_q.put((po, "running", "Logging into Amazon Vendor Central…"))
                    open_amazon_ship_builder(driver)
                    _amz_login(driver, lm)
                    login_done = True
                else:
                    open_amazon_ship_builder(driver)

                status_q.put((po, "running", "Running Shipment Builder…"))
                amazon_shipment_builder(driver, username, excel_path)
                status_q.put((po, "complete", "Shipment submitted ✓"))

            except Exception as e:
                status_q.put((po, "failed", f"Error: {str(e)[:90]}"))

        # ── Pick tickets ───────────────────────────────────────────────────
        status_q.put(("__ALL__", "info", "Printing pick tickets from NetSuite…"))
        try:
            driver.get(NETSUITE_PICK_TICKETS)
            driver.maximize_window()
            _ns_login(driver, lm)
            time.sleep(2)

            search_field = wait_for_element_to_be_clickable(
                driver, By.XPATH,
                '//input[@id="item_Transaction_NAME_display"]'
            )
            if search_field:
                search_field.click()
                search_field.send_keys(Keys.CONTROL, "a")
                search_field.send_keys(Keys.DELETE)
                search_field.send_keys("11187 AMAZON.COM/SEATTLE WA")
                search_field.send_keys(Keys.ARROW_DOWN)
                search_field.send_keys(Keys.ENTER)
                time.sleep(5)

            po_list_from_sheets = [
                re.search(r"Pack Sheet - (.+)\.pdf", os.path.basename(f)).group(1)
                for f in glob.glob(os.path.join(AMAZON_DAILY_PACK_SHEETS, "Pack Sheet - *.pdf"))
            ]

            for po in po_list_from_sheets:
                try:
                    log_message(f"Looking for pick ticket row: {po}")
                    row = wait_for_element(
                        driver, By.XPATH,
                        f'//tr[contains(@class,"uir-list-row-tr") and '
                        f'.//td[normalize-space(text())="{po}"]]',
                        timeout=5,
                    )
                    if row:
                        log_message(f"Found row for PO: {po}")
                        checkbox_element = row.find_element(By.XPATH, './/td[1]//span/input[@type="checkbox"]')
                        if checkbox_element and checkbox_element.is_displayed():
                            checkbox_element.click()
                            log_message(f"Clicked checkbox for PO: {po}")
                            status_q.put(("__LOG__", "info", f"Checked pick ticket for PO {po}"))
                        else:
                            log_message(f"Checkbox element not found or not visible for PO: {po}")
                    else:
                        log_message(f"Row not found for PO {po}")
                except Exception as e_pt:
                    log_message(f"Error finding pick ticket row for {po}: {e_pt}")

            time.sleep(2)

            log_message("Clicking the Print Pick Tickets button")
            # Use the button id directly — much more reliable than an absolute XPath
            final_btn = wait_for_element_to_be_clickable(
                driver, By.ID, "nl_print", timeout=10,
            )
            if final_btn:
                driver.execute_script("arguments[0].click();", final_btn)
                log_message("Clicked Print Pick Tickets button")
            else:
                log_message("Print Pick Tickets button not found")

            # Wait up to 60s for PICKINGTICKET file
            log_message("Waiting for PICKINGTICKET file in Downloads…")
            timeout = 60
            start = time.time()
            while not any(f.startswith("PICKINGTICKET") for f in os.listdir(DOWNLOAD_FOLDER)):
                if time.time() - start > timeout:
                    log_message("Timed out waiting for PICKINGTICKET file")
                    break
                log_message("PICKINGTICKET not found yet, waiting…")
                time.sleep(5)

            for f in os.listdir(DOWNLOAD_FOLDER):
                if f.startswith("PICKINGTICKET"):
                    time.sleep(2)
                    shutil.move(
                        os.path.join(DOWNLOAD_FOLDER, f),
                        os.path.join(AMAZON_DAILY_FOLDER, "Pick Tickets.pdf"),
                    )
                    log_message("Pick Tickets.pdf saved")
                    status_q.put(("__LOG__", "ok", "Pick Tickets saved."))
                    break

        except Exception as e:
            status_q.put(("__ALL__", "warn", f"Pick tickets error (non-fatal): {e}"))

        # ── Merge PDFs ─────────────────────────────────────────────────────
        for folder, name in [
            (AMAZON_DAILY_PACK_SHEETS,   "Merged Pack Sheets.pdf"),
            (AMAZON_DAILY_CARTON_LABELS, "Merged Carton Labels.pdf"),
            (AMAZON_DAILY_PALLET_LABELS, "Merged Pallet Labels.pdf"),
        ]:
            status_q.put(("__ALL__", "info", f"Merging {name}…"))
            merge_pdfs(folder, os.path.join(AMAZON_DAILY_FOLDER, name))

        # ── Zip ────────────────────────────────────────────────────────────
        today_label = datetime.now().strftime("%d-%b")
        zip_name    = f"From {today_label}.zip"
        zip_path    = os.path.join(AMAZON_DAILY_FOLDER, zip_name)
        temp_zip    = os.path.join(AMAZON_WEEKLY_FOLDER, zip_name)
        status_q.put(("__ALL__", "info", f"Creating zip: {zip_name}…"))
        with zipfile.ZipFile(temp_zip, "w", zipfile.ZIP_DEFLATED) as zf:
            zipdir(AMAZON_DAILY_FOLDER, zf, "Label Upload", "Canada")
        shutil.move(temp_zip, zip_path)

        # ── Email ──────────────────────────────────────────────────────────
        status_q.put(("__ALL__", "info", "Preparing email…"))
        csv_path = os.path.join(AMAZON_DAILY_FOLDER, "selected_dates.csv")
        data = []
        if os.path.exists(csv_path):
            with open(csv_path) as f:
                for row in csv.reader(f):
                    so, po, date_str, ship_type = row
                    data.append((datetime.strptime(date_str, "%B %d, %Y"), so, po, ship_type))
            data.sort()

        rows_html = "".join(
            f"<tr>"
            f"<td style='width:2in;font-family:Helvetica;font-size:12pt;text-align:center'>{so}</td>"
            f"<td style='width:2in;font-family:Helvetica;font-size:12pt;text-align:center'>{po}</td>"
            f"<td style='width:2in;font-family:Helvetica;font-size:12pt;text-align:center'>{dt.strftime('%B %d, %Y')}</td>"
            f"<td style='width:2in;font-family:Helvetica;font-size:12pt;text-align:center'>{stype}</td>"
            f"</tr>"
            for dt, so, po, stype in data
        )
        table_html = (
            "<table border='1' style='border-collapse:collapse'>"
            "<tr>"
            "<th style='width:2in;font-family:Helvetica;font-size:12pt;text-align:center'>SO #</th>"
            "<th style='width:2in;font-family:Helvetica;font-size:12pt;text-align:center'>PO #</th>"
            "<th style='width:2in;font-family:Helvetica;font-size:12pt;text-align:center'>Scheduled Ship Date</th>"
            "<th style='width:2in;font-family:Helvetica;font-size:12pt;text-align:center'>Shipment Type</th>"
            "</tr>"
            + rows_html + "</table>"
        )

        monday_label = (datetime.now() - timedelta(days=datetime.now().weekday())).strftime("%d-%b")
        subject = f"Amazon US Orders - Week of {monday_label} - From {today_label}"
        body    = (
            "<p style='font-family:Helvetica,Arial,sans-serif;font-size:12pt'>Team,</p>"
            "<p style='font-family:Helvetica,Arial,sans-serif;font-size:12pt'>"
            "Here is the paperwork for the following orders:</p>"
            f"{table_html}<br><br>"
            "<p style='font-family:Helvetica,Arial,sans-serif;font-size:12pt'>Thank you,</p>"
        )
        recipients = [
            "sales1@your-company.com",
            "ops1@your-company.com",
            "sales2@your-company.com",
            "mgmt1@your-company.com",
            "ops2@your-company.com",
            "mgmt2@your-company.com",
        ]
        send_email_with_outlook(subject, body, recipients, [zip_path])
        status_q.put(("__ALL__", "ok", "Email draft ready — review and send in Outlook."))

        # ── Save log ───────────────────────────────────────────────────────
        log_name = f"Order Automation Log - {datetime.now().strftime('%Y - %m (%b) - %d')}.txt"
        with open(os.path.join(AMAZON_DAILY_FOLDER, log_name), "w") as lf:
            lf.write("\n".join(_automation_log))

        driver.quit()

        # ── Resume OneDrive ───────────────────────────────────────────────
        od_ok, od_msg = _onedrive_resume()
        status_q.put(("__LOG__", "ok" if od_ok else "warn", f"OneDrive: {od_msg}"))
        status_q.put(("__ONEDRIVE__", "resumed", od_msg))

        status_q.put(("__DONE__", "ok", "All orders processed successfully."))

    except Exception as e:
        # Always try to resume OneDrive even on failure
        try:
            od_ok, od_msg = _onedrive_resume()
            status_q.put(("__LOG__", "ok" if od_ok else "warn", f"OneDrive: {od_msg}"))
            status_q.put(("__ONEDRIVE__", "resumed", od_msg))
        except Exception:
            pass
        status_q.put(("__DONE__", "error", f"Fatal error: {e}"))


# ════════════════════════════════════════════════════════════════════════════
# PAGE UI
# ════════════════════════════════════════════════════════════════════════════
st.title("📦 Generate Amazon Paperwork")
st.divider()

phase = st.session_state.ppr_phase

left_col, right_col = st.columns([1.4, 1])

# ── Right: order status cards (always visible once POs are known) ────────
with right_col:
    st.markdown("### Order Status")

    # OneDrive status badge (only shown once automation has started)
    if st.session_state.ppr_phase not in ("idle", "netsuite_export",
                                           "cross_ref_check", "cross_ref_fix",
                                           "split_csv"):
        if st.session_state.ppr_onedrive_paused:
            st.markdown(
                '<div style="display:inline-flex;align-items:center;gap:8px;'
                'background:#3d2e00;border:1px solid #7a6000;border-radius:6px;'
                'padding:6px 14px;margin-bottom:12px;font-size:0.82rem;color:#f5c542">'
                '⏸ OneDrive paused</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div style="display:inline-flex;align-items:center;gap:8px;'
                'background:#0f3d26;border:1px solid #1a6040;border-radius:6px;'
                'padding:6px 14px;margin-bottom:12px;font-size:0.82rem;color:#3ee87a">'
                '▶ OneDrive running</div>',
                unsafe_allow_html=True,
            )
    if not st.session_state.ppr_po_list:
        st.markdown(
            '<div style="background:#141720;border:1px solid #1e2235;border-radius:8px;'
            'padding:28px;text-align:center;color:#50566a">'
            'Orders appear here after the<br>worksheet is downloaded &amp; split.'
            '</div>',
            unsafe_allow_html=True,
        )
    else:
        total    = len(st.session_state.ppr_po_list)
        complete = sum(1 for v in st.session_state.ppr_order_status.values()
                       if v["status"] == "complete")
        if phase == "processing" and total:
            st.progress(complete / total, text=f"{complete}/{total} orders complete")
        st.markdown(_order_cards(), unsafe_allow_html=True)

# ── Left: step UI ────────────────────────────────────────────────────────
with left_col:

    # ── IDLE / PREFLIGHT ────────────────────────────────────────────────
    if phase == "idle":
        st.markdown("#### Pre-flight Checklist")
        poa_ok = st.checkbox("All POAs have been sent and show 'Confirmed' in Vendor Central")
        issues_ok = st.checkbox("All inventory issues showing '0' in NetSuite")

        st.info(
            "☁️ **OneDrive will be paused automatically** when the automation starts "
            "and restarted automatically when it finishes.  \n"
            "🔗 **Cross-reference check runs automatically** after the worksheet downloads.",
            icon="ℹ️",
        )

        all_ok = poa_ok and issues_ok
        if not all_ok:
            st.caption("⚠️ Check all items above to enable the Start button.")

        if st.button("🚀 Start Automation", disabled=not all_ok, type="primary"):
            st.session_state.ppr_phase = "netsuite_export"
            _log("Pre-flight confirmed. Waiting for NetSuite CSV export.", "ok")
            st.rerun()

    # ── NETSUITE EXPORT ─────────────────────────────────────────────────
    elif phase == "netsuite_export":
        st.markdown("#### Step 1 — NetSuite Export")

        # ── Spin up the export thread exactly once ──────────────────────
        if not st.session_state.ppr_export_started:
            st.session_state.ppr_export_started = True
            eq = queue.Queue()
            st.session_state.ppr_export_q = eq
            t = threading.Thread(
                target=_run_netsuite_export,
                args=(eq,),
                daemon=True,
            )
            t.start()
            st.session_state.ppr_export_thread = t
            _log("Chrome launched — logging into NetSuite…", "info")

        # ── Drain export queue ──────────────────────────────────────────
        eq = st.session_state.ppr_export_q
        if eq:
            while not eq.empty():
                key, level, msg = eq.get_nowait()
                if key == "__DONE__":
                    if level == "ok":
                        # msg is the CSV file path
                        found_path = msg
                        try:
                            df = pd.read_csv(found_path, dtype={"Item": str})
                            st.session_state.ppr_combined_path = found_path
                            st.session_state.ppr_combined_df   = df
                            pos = (df["PO"].astype(str)
                                   .str.replace("PO:", "").str.strip().unique())
                            _log(f"Combined CSV ready: {os.path.basename(found_path)} "
                                 f"({len(df)} rows, {len(pos)} PO(s))", "ok")
                            st.session_state.ppr_phase = "cross_ref_check"
                        except Exception as e:
                            _log(f"Could not read CSV: {e}", "error")
                    else:
                        _log(msg, "error")
                else:
                    _log(msg, level)

        # ── If already advanced (CSV loaded, phase flipped) rerun now ──
        if st.session_state.ppr_phase != "netsuite_export":
            st.rerun()

        # ── Show live status while waiting ──────────────────────────────
        export_thread = st.session_state.ppr_export_thread
        is_running = export_thread is not None and export_thread.is_alive()

        if is_running:
            last_msg = ""
            for e in reversed(st.session_state.ppr_log):
                if e["level"] in ("info", "ok"):
                    last_msg = e["msg"]
                    break
            st.info(f"🌐 **Chrome is open and working…**\n\n{last_msg}", icon="🔄")
            # Auto-refresh every second while the thread is alive
            time.sleep(1)
            st.rerun()
        else:
            # Thread finished but phase hasn't flipped — something went wrong
            st.error(
                "The NetSuite export did not complete automatically.\n\n"
                "You can trigger it manually: log into NetSuite, open the pack worksheet "
                "search, and click **Export - CSV**. Once the file appears in your "
                f"Downloads folder (`{DOWNLOAD_FOLDER}`), click **Check Downloads**."
            )
            if st.button("🔄 Check Downloads"):
                found = _find_combined_csv()
                if found:
                    try:
                        df = pd.read_csv(found, dtype={"Item": str})
                        st.session_state.ppr_combined_path = found
                        st.session_state.ppr_combined_df   = df
                        _log(f"CSV loaded manually: {found}", "ok")
                        st.session_state.ppr_phase = "cross_ref_check"
                        st.rerun()
                    except Exception as e:
                        st.error(str(e)); _log(str(e), "error")
                else:
                    st.warning("Still not found — try again in a moment.")
                    _log("Manual check: CSV not found yet.", "warn")

    # ── CROSS-REF CHECK (auto-runs, no button needed) ───────────────────
    elif phase == "cross_ref_check":
        st.markdown("#### Step 2 — Cross-Reference Check")
        st.caption(
            "Checking every `Item` in the combined worksheet against "
            "`Amazon Item Cross References.csv` before splitting into per-PO files."
        )

        with st.spinner("Checking cross-reference…"):
            try:
                cross_df, missing = _do_cross_ref_check(
                    st.session_state.ppr_combined_df
                )
                st.session_state.ppr_cross_ref_df = cross_df
                st.session_state.ppr_missing      = missing

                if missing:
                    _log(f"Cross-ref: {len(missing)} missing item(s): {missing}", "warn")
                    st.session_state.ppr_phase = "cross_ref_fix"
                else:
                    _log("Cross-ref: all items present ✓", "ok")
                    st.session_state.ppr_phase = "split_csv"
                st.rerun()

            except FileNotFoundError:
                st.error(
                    f"Cross-reference file not found:\n\n`{AMAZON_PW_ITEM_CROSS_REF}`\n\n"
                    "Check that the Pack Worksheet folder path in secrets.toml is correct."
                )
                _log("Cross-ref file not found.", "error")
            except Exception as e:
                st.error(str(e)); _log(str(e), "error")

    # ── CROSS-REF FIX ───────────────────────────────────────────────────
    elif phase == "cross_ref_fix":
        missing = st.session_state.ppr_missing
        st.markdown(f"#### ⚠️ {len(missing)} Item(s) Missing from Cross-Reference")
        st.warning(
            "These items are in the worksheet but not in `Amazon Item Cross References.csv`. "
            "Enter the **BPN** and **UPC** for each, then click **Save & Continue**. "
            "The file will be updated before splitting begins."
        )

        for item in missing:
            if item not in st.session_state.ppr_missing_edits:
                st.session_state.ppr_missing_edits[item] = {"BPN": "", "UPC": ""}

        # Build item → first PO lookup from the combined DataFrame
        combined_df = st.session_state.ppr_combined_df
        item_to_po: dict = (
            combined_df.drop_duplicates(subset=["Item"])
            .set_index(combined_df["Item"].astype(str).str.strip())["PO"]
            .apply(lambda p: str(p).replace("PO:", "").strip())
            .to_dict()
        )

        VC_PO_URL = "https://vendorcentral.amazon.com/po/vendor/members/po-mgmt/order?poId={po}"

        c0, c1, c2, c3 = st.columns([1.4, 1.8, 1.8, 1.2])
        c0.markdown("**Item #**")
        c1.markdown("**BPN / ASIN**")
        c2.markdown("**UPC**")
        c3.markdown("**Vendor Central**")

        for item in missing:
            r0, r1, r2, r3 = st.columns([1.4, 1.8, 1.8, 1.2])
            r0.markdown(f"`{item}`")
            bpn = r1.text_input("BPN", key=f"ppr_bpn_{item}",
                                 label_visibility="collapsed", placeholder="BPN / ASIN")
            upc = r2.text_input("UPC", key=f"ppr_upc_{item}",
                                 label_visibility="collapsed", placeholder="UPC barcode")
            st.session_state.ppr_missing_edits[item] = {"BPN": bpn, "UPC": upc}

            first_po = item_to_po.get(item, "")
            if first_po:
                r3.markdown(
                    f'<a href="{VC_PO_URL.format(po=first_po)}" target="_blank" '
                    f'style="font-size:0.82rem;color:#40aaff;text-decoration:none;" '
                    f'title="Open PO {first_po} in Vendor Central">'
                    f'🔗 PO {first_po}</a>',
                    unsafe_allow_html=True,
                )
            else:
                r3.markdown('<span style="color:#505568;font-size:0.8rem">—</span>',
                            unsafe_allow_html=True)

        st.markdown("")
        if st.button("💾 Save & Continue", type="primary"):
            new_rows = [
                {
                    "Item":     item,
                    "Customer": "11187 AMAZON.COM/SEATTLE WA",
                    "BPN":      st.session_state.ppr_missing_edits[item].get("BPN", ""),
                    "UPC":      st.session_state.ppr_missing_edits[item].get("UPC", ""),
                }
                for item in missing
            ]
            try:
                updated = _save_cross_ref(st.session_state.ppr_cross_ref_df, new_rows)
                st.session_state.ppr_cross_ref_df = updated
                st.session_state.ppr_missing      = []
                st.session_state.ppr_missing_edits = {}
                st.session_state.ppr_phase         = "split_csv"
                _log(f"Cross-ref updated: {len(new_rows)} item(s) saved.", "ok")
                st.rerun()
            except Exception as e:
                st.error(f"Failed to save: {e}"); _log(str(e), "error")

    # ── SPLIT CSV (auto) ─────────────────────────────────────────────────
    elif phase == "split_csv":
        st.markdown("#### Step 3 — Splitting into Per-PO Files")
        with st.spinner("All items verified ✓ — splitting…"):
            try:
                po_list = _split_combined(st.session_state.ppr_combined_df)

                # Delete combined file (mirrors original script)
                cp = st.session_state.ppr_combined_path
                if cp and os.path.exists(cp):
                    os.remove(cp)
                    _log(f"Deleted combined file: {os.path.basename(cp)}", "info")

                st.session_state.ppr_po_list = po_list
                for po in po_list:
                    st.session_state.ppr_order_status[po] = {
                        "status": "pending", "step": "Queued",
                        "start": None, "end": None,
                    }
                st.session_state.ppr_phase   = "processing"
                st.session_state.ppr_running = True
                _log(f"Split complete — {len(po_list)} PO(s): {po_list}", "ok")
                st.rerun()
            except Exception as e:
                st.error(f"Split error: {e}"); _log(str(e), "error")

    # ── PROCESSING ───────────────────────────────────────────────────────
    elif phase == "processing":
        st.markdown("#### Step 4 — Processing Orders")

        # Spin up background thread exactly once using a boolean flag,
        # NOT is_alive() — is_alive() is unreliable across fast st.rerun() cycles
        # and would respawn the thread (and re-run makedirs → opens Explorer window)
        if not st.session_state.ppr_run_started:
            st.session_state.ppr_run_started = True
            q = queue.Queue()
            st.session_state.ppr_status_q = q
            t = threading.Thread(
                target=_run_automation,
                args=(st.session_state.ppr_po_list, q),
                daemon=True,
            )
            t.start()
            st.session_state.ppr_run_thread = t

        # Drain queue
        q = st.session_state.ppr_status_q
        if q:
            while not q.empty():
                key, status_type, msg = q.get_nowait()
                if key == "__DONE__":
                    st.session_state.ppr_phase   = "done"
                    st.session_state.ppr_running = False
                    _log(msg, "ok" if status_type == "ok" else "error")
                elif key == "__ONEDRIVE__":
                    st.session_state.ppr_onedrive_paused = (status_type == "paused")
                    _log(msg, "ok")
                elif key in ("__ALL__", "__LOG__"):
                    _log(msg, status_type)
                else:
                    _update_order(key, status_type, msg)

        if st.session_state.ppr_running:
            time.sleep(0.8)
            st.rerun()

    # ── DONE ─────────────────────────────────────────────────────────────
    elif phase == "done":
        completed = [p for p, v in st.session_state.ppr_order_status.items()
                     if v["status"] == "complete"]
        failed    = [p for p, v in st.session_state.ppr_order_status.items()
                     if v["status"] == "failed"]

        st.success(f"🎉 Done — {len(completed)} order(s) complete, {len(failed)} failed.")
        if failed:
            st.error("Failed POs: " + ", ".join(f"`{p}`" for p in failed))
        st.info(f"📁 Output: `{AMAZON_DAILY_FOLDER}`", icon="📁")

        if st.button("🔄 Run Again (new day)", type="primary"):
            for k in [
                "ppr_phase", "ppr_log", "ppr_combined_path", "ppr_combined_df",
                "ppr_cross_ref_df", "ppr_missing", "ppr_missing_edits",
                "ppr_po_list", "ppr_order_status", "ppr_running",
                "ppr_status_q", "ppr_run_thread", "ppr_export_thread",
                "ppr_export_q", "ppr_export_started", "ppr_run_started", "ppr_onedrive_paused",
            ]:
                st.session_state.pop(k, None)
            _init()
            st.rerun()

# ── Activity Log (always visible at bottom) ──────────────────────────────
st.divider()
with st.expander("📋 Activity Log", expanded=(phase in ("processing", "done"))):
    log_placeholder = st.empty()
    _render_log(log_placeholder)

    if st.session_state.ppr_log:
        log_text = "\n".join(
            f"[{e['ts']}] [{e['level'].upper()}] {e['msg']}"
            for e in st.session_state.ppr_log
        )
        st.download_button(
            "⬇ Download Log",
            data=log_text,
            file_name=f"amz_paperwork_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            mime="text/plain",
        )
